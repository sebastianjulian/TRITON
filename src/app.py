# # from flask import Flask, request, jsonify, render_template_string
# # import threading

# # app = Flask(__name__)
# # latest_data = {"timestamp": "No data received", "data": {}}
# # lock = threading.Lock()

# # # HTML template for the dashboard
# # HTML_TEMPLATE = """
# # <!DOCTYPE html>
# # <html>
# # <head>
# #     <title>CanSat Sensor Dashboard</title>
# #     <meta http-equiv="refresh" content="2">
# #     <style>
# #         body { font-family: Arial, sans-serif; background: #f4f4f4; padding: 2em; }
# #         h1 { color: #333; }
# #         table { border-collapse: collapse; width: 60%; }
# #         th, td { border: 1px solid #888; padding: 8px 12px; text-align: left; }
# #         th { background-color: #eee; }
# #     </style>
# # </head>
# # <body>
# #     <h1>CanSat Live Sensor Data</h1>
# #     <p><strong>Last Update:</strong> {{ timestamp }}</p>
# #     <table>
# #         <tr><th>Sensor</th><th>Value</th></tr>
# #         {% for key, value in data.items() %}
# #         <tr><td>{{ key }}</td><td>{{ value }}</td></tr>
# #         {% endfor %}
# #     </table>
# # </body>
# # </html>
# # """

# # @app.route("/", methods=["GET"])
# # def dashboard():
# #     with lock:
# #         return render_template_string(HTML_TEMPLATE, timestamp=latest_data["timestamp"], data=latest_data["data"])

# # @app.route("/update", methods=["POST"])
# # def update_data():
# #     global latest_data
# #     new_data = request.get_json()
# #     if new_data:
# #         with lock:
# #             latest_data = new_data
# #     return jsonify(status="ok")

# # if __name__ == "__main__":
# #     app.run(host="0.0.0.0", port=5000)


# from flask import Flask, request, jsonify, render_template
# from collections import defaultdict
# import time

# app = Flask(__name__)

# # Store all sensor data (x = elapsed, y = value)
# data_history = defaultdict(lambda: {"x": [], "y": []})
# MAX_POINTS = 300  # number of points shown per graph

# @app.route("/", methods=["GET"])
# def dashboard():
#     return render_template("dashboard.html")

# @app.route("/update", methods=["POST"])
# def update_data():
#     content = request.json
#     elapsed = float(content["data"]["Elapsed [s]"])
    
#     for key, value in content["data"].items():
#         if key == "Elapsed [s]":
#             continue
#         try:
#             y_val = float(value)
#         except:
#             continue  # skip non-numeric
#         d = data_history[key]
#         d["x"].append(elapsed)
#         d["y"].append(y_val)
#         if len(d["x"]) > MAX_POINTS:
#             d["x"] = d["x"][-MAX_POINTS:]
#             d["y"] = d["y"][-MAX_POINTS:]
#     return "OK"

# @app.route("/data", methods=["GET"])
# def get_data():
#     return jsonify(data_history)

# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=5000)


import os
import signal
import subprocess
import atexit
import platform
import random
import time
import json
import glob as glob_module
import io
import csv
import threading
from datetime import datetime

from flask import Flask, request, jsonify, render_template, send_from_directory, Response

# Configuration directory setup
CONFIG_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'config'))
CONFIG_FILE = os.path.join(CONFIG_DIR, 'triton_config.json')

# Default configuration values
DEFAULT_CONFIG = {
    "sea_level_pressure": 993.9,
    "transmission_thresholds": {
        "temp_bme280": 0.25,
        "humidity": 1.0,
        "pressure": 0.5,
        "altitude": 0.5,
        "acceleration": 0.25,
        "gyroscope": 5.0,
        "temp_mpu": 0.25
    },
    "update_frequency": 1.0,
    "history_length": 300
}

def load_config():
    """Load configuration from file or return defaults"""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"[WARN] Could not load config: {e}")
    return DEFAULT_CONFIG.copy()

def save_config(config):
    """Save configuration to file"""
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4)

def get_profile_path(name):
    """Get the path for a configuration profile"""
    # Sanitize profile name
    safe_name = "".join(c for c in name if c.isalnum() or c in ('-', '_')).strip()
    if not safe_name:
        raise ValueError("Invalid profile name")
    return os.path.join(CONFIG_DIR, f'profile_{safe_name}.json')

def list_profiles():
    """List all saved configuration profiles"""
    profiles = []
    if os.path.exists(CONFIG_DIR):
        for f in os.listdir(CONFIG_DIR):
            if f.startswith('profile_') and f.endswith('.json'):
                name = f[8:-5]  # Remove 'profile_' prefix and '.json' suffix
                profiles.append(name)
    return profiles

def kill_port(port=5000):
    try:
        if platform.system() == "Windows":
            # Windows command to find processes using the port
            output = subprocess.check_output(f'netstat -ano | findstr :{port}', shell=True).decode().strip()
            if output:
                lines = output.splitlines()
                for line in lines:
                    if 'LISTENING' in line:
                        pid = line.split()[-1]
                        try:
                            subprocess.run(f'taskkill /F /PID {pid}', shell=True, check=True)
                            print(f"[INFO] Killed process on port {port} (PID: {pid})")
                        except subprocess.CalledProcessError as e:
                            print(f"[WARN] Couldn't kill PID {pid}: {e}")
        else:
            # Unix/Linux command
            output = subprocess.check_output(f"lsof -ti:{port}", shell=True).decode().split()
            for pid in output:
                try:
                    os.kill(int(pid), signal.SIGKILL)
                    print(f"[INFO] Killed process on port {port} (PID: {pid})")
                except Exception as e:
                    print(f"[WARN] Couldn't kill PID {pid}: {e}")
    except subprocess.CalledProcessError:
        # No process found using the port
        pass
    except Exception as e:
        print(f"[WARN] Port cleanup failed: {e}")

kill_port(5000)

app = Flask(__name__, template_folder='templates')

# Directory setup for downloads
LOG_DIR = os.path.abspath("logs")
ARCHIVE_DIR = os.path.join(LOG_DIR, "previous_data")
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)

latest_data = {"timestamp": "", "data": {}}
history = {"Elapsed [s]": []}  # For graphing
for key in [
    "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]",
    "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
    "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]",
    "Temp_MPU [°C]"
]:
    history[key] = []

# ==================== Recording State ====================
RECORDINGS_DIR = os.path.join(LOG_DIR, "recordings")
os.makedirs(RECORDINGS_DIR, exist_ok=True)

recording_state = {
    "is_recording": False,
    "filename": None,
    "filepath": None,
    "start_time": None,
    "point_count": 0
}
recording_lock = threading.Lock()

@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")

@app.route("/dashboard", methods=["GET"])
def dashboard():
    return render_template("dashboard.html")


@app.route("/update", methods=["POST"])
def update():
    global latest_data, history
    content = request.json
    latest_data = content

    elapsed = float(content["data"]["Elapsed [s]"])
    history["Elapsed [s]"].append(elapsed)

    for key in history:
        if key != "Elapsed [s]":
            try:
                history[key].append(float(content["data"][key]))
            except:
                history[key].append(None)

    # Write to recording if active
    append_recording_data(content["data"], content.get("timestamp", datetime.now().isoformat()))

    return jsonify(status="ok")


@app.route("/data", methods=["GET"])
def get_data():
    return jsonify({
        "timestamp": latest_data["timestamp"],
        "history": history
    })

@app.route("/generate_random_data", methods=["POST"])
def generate_random_data():
    global latest_data, history
    
    # Get current elapsed time or start from 0
    current_elapsed = history["Elapsed [s]"][-1] if history["Elapsed [s]"] else 0
    new_elapsed = current_elapsed + random.uniform(1.0, 3.0)  # 1-3 second increment
    
    # Generate realistic sensor data
    random_data = {
        "Elapsed [s]": round(new_elapsed, 2),
        "Temp_BME280 [°C]": round(random.uniform(15.0, 35.0) + random.gauss(0, 0.5), 2),
        "Hum [%]": round(random.uniform(30.0, 80.0) + random.gauss(0, 2.0), 1),
        "Press [hPa]": round(random.uniform(980.0, 1020.0) + random.gauss(0, 1.0), 1),
        "Alt [m]": round(random.uniform(-50.0, 200.0) + random.gauss(0, 5.0), 1),
        "Acc x [m/s²]": round(random.gauss(0, 0.3), 3),
        "Acc y [m/s²]": round(random.gauss(0, 0.3), 3),
        "Acc z [m/s²]": round(random.uniform(9.5, 10.2) + random.gauss(0, 0.2), 3),  # Gravity + noise
        "Gyro x [°/s]": round(random.gauss(0, 10.0), 2),
        "Gyro y [°/s]": round(random.gauss(0, 10.0), 2),
        "Gyro z [°/s]": round(random.gauss(0, 10.0), 2),
        "Temp_MPU [°C]": round(random.uniform(20.0, 40.0) + random.gauss(0, 0.5), 2)
    }
    
    # Create properly formatted data structure
    generated_payload = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "data": random_data
    }
    
    # Update global data structures (same as /update endpoint)
    latest_data = generated_payload
    history["Elapsed [s]"].append(new_elapsed)
    
    for key in history:
        if key != "Elapsed [s]":
            try:
                history[key].append(float(random_data[key]))
            except:
                history[key].append(None)

    # Write to recording if active
    append_recording_data(random_data, generated_payload["timestamp"])

    return jsonify({"status": "success", "message": "Random data generated"})

@app.route("/save_session_csv", methods=["POST"])
def save_session_csv():
    """Save current session data to a proper CSV file"""
    global history

    if not history.get("Elapsed [s]") or len(history["Elapsed [s]"]) == 0:
        return jsonify({"status": "error", "message": "No data to save"}), 400

    try:
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"sensor_data_{timestamp}.csv"
        filepath = os.path.join(LOG_DIR, filename)

        # Move any existing files to archive
        import glob
        import shutil
        for file in glob.glob(os.path.join(LOG_DIR, "sensor_data_*.csv")):
            shutil.move(file, ARCHIVE_DIR)

        # Write CSV with proper comma-delimited format
        headers = ["MET [s]"] + [
            "Elapsed [s]", "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]",
            "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
            "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]", "Temp_MPU [°C]"
        ]

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(",".join(headers) + "\n")

            num_rows = len(history.get("Elapsed [s]", []))
            for i in range(num_rows):
                row = []
                # MET timestamp (use elapsed time as MET)
                elapsed = history.get("Elapsed [s]", [0])[i] if i < len(history.get("Elapsed [s]", [])) else 0
                row.append(f"{elapsed:.3f}")

                # Add all other columns
                for key in headers[1:]:
                    values = history.get(key, [])
                    if i < len(values) and values[i] is not None:
                        row.append(f"{values[i]:.3f}" if isinstance(values[i], (int, float)) else str(values[i]))
                    else:
                        row.append("")

                f.write(",".join(row) + "\n")

        return jsonify({
            "status": "success",
            "message": f"Saved {num_rows} rows to {filename}",
            "filename": filename,
            "rows": num_rows
        })

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/clear_test_data", methods=["POST"])
def clear_test_data():
    global latest_data, history
    
    # Clear all data
    latest_data = {"timestamp": "", "data": {}}
    history = {"Elapsed [s]": []}  # Reset to empty arrays
    for key in [
        "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]",
        "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
        "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]",
        "Temp_MPU [°C]"
    ]:
        history[key] = []
    
    return jsonify({"status": "success", "message": "Test data cleared"})

@app.route('/download/latest')
def download_latest():
    files = sorted([f for f in os.listdir(LOG_DIR) if f.startswith("sensor_data_")], reverse=True)
    if files:
        return send_from_directory(LOG_DIR, files[0], as_attachment=True)
    return "No log files found", 404

@app.route('/download/list')
def list_logs():
    """List all downloadable files from archive and recordings directories"""
    files = []

    # Get archived logs
    try:
        for f in sorted(os.listdir(ARCHIVE_DIR)):
            if f.endswith(".csv"):
                files.append({"filename": f, "path": "archive", "display": f"[Archive] {f}"})
    except Exception as e:
        print(f"[WARN] Error listing archive: {e}")

    # Get recordings
    try:
        for f in sorted(os.listdir(RECORDINGS_DIR)):
            if f.endswith(".csv"):
                files.append({"filename": f, "path": "recording", "display": f"[Recording] {f}"})
    except Exception as e:
        print(f"[WARN] Error listing recordings: {e}")

    # Sort by filename (which includes timestamp)
    files.sort(key=lambda x: x["filename"], reverse=True)
    return jsonify(files)


@app.route('/download/archive/<filename>')
def download_archive(filename):
    """Download from archive or recordings directory"""
    safe_filename = os.path.basename(filename)

    # Check recordings first (newer feature)
    rec_path = os.path.join(RECORDINGS_DIR, safe_filename)
    if os.path.isfile(rec_path):
        return send_from_directory(RECORDINGS_DIR, safe_filename, as_attachment=True)

    # Check archive
    arch_path = os.path.join(ARCHIVE_DIR, safe_filename)
    if os.path.isfile(arch_path):
        return send_from_directory(ARCHIVE_DIR, safe_filename, as_attachment=True)

    return "File not found", 404


# ==================== Recording API Endpoints ====================

def write_recording_header(filepath):
    """Write CSV header to recording file"""
    headers = ["Timestamp", "Elapsed [s]", "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]",
               "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
               "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]", "Temp_MPU [°C]"]
    with open(filepath, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)


def append_recording_data(data_dict, timestamp):
    """Append a data point to the active recording file"""
    global recording_state

    with recording_lock:
        if not recording_state["is_recording"] or not recording_state["filepath"]:
            return

        try:
            row = [
                timestamp,
                data_dict.get("Elapsed [s]", ""),
                data_dict.get("Temp_BME280 [°C]", ""),
                data_dict.get("Hum [%]", ""),
                data_dict.get("Press [hPa]", ""),
                data_dict.get("Alt [m]", ""),
                data_dict.get("Acc x [m/s²]", ""),
                data_dict.get("Acc y [m/s²]", ""),
                data_dict.get("Acc z [m/s²]", ""),
                data_dict.get("Gyro x [°/s]", ""),
                data_dict.get("Gyro y [°/s]", ""),
                data_dict.get("Gyro z [°/s]", ""),
                data_dict.get("Temp_MPU [°C]", "")
            ]
            with open(recording_state["filepath"], 'a', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(row)
            recording_state["point_count"] += 1
        except Exception as e:
            print(f"[RECORDING] Error appending data: {e}")


@app.route('/recording/start', methods=['POST'])
def start_recording():
    """Start a new recording session"""
    global recording_state

    with recording_lock:
        if recording_state["is_recording"]:
            return jsonify({
                "status": "error",
                "message": "Recording already in progress",
                "filename": recording_state["filename"]
            }), 400

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"recording_{timestamp}.csv"
        filepath = os.path.join(RECORDINGS_DIR, filename)

        # Write header
        write_recording_header(filepath)

        # Update state
        recording_state["is_recording"] = True
        recording_state["filename"] = filename
        recording_state["filepath"] = filepath
        recording_state["start_time"] = datetime.now().isoformat()
        recording_state["point_count"] = 0

        print(f"[RECORDING] Started recording to {filename}")

        return jsonify({
            "status": "success",
            "message": "Recording started",
            "filename": filename,
            "start_time": recording_state["start_time"]
        })


@app.route('/recording/stop', methods=['POST'])
def stop_recording():
    """Stop the current recording session"""
    global recording_state

    with recording_lock:
        if not recording_state["is_recording"]:
            return jsonify({
                "status": "error",
                "message": "No recording in progress"
            }), 400

        filename = recording_state["filename"]
        point_count = recording_state["point_count"]
        start_time = recording_state["start_time"]

        # Reset state
        recording_state["is_recording"] = False
        recording_state["filename"] = None
        recording_state["filepath"] = None
        recording_state["start_time"] = None
        recording_state["point_count"] = 0

        print(f"[RECORDING] Stopped recording. Saved {point_count} points to {filename}")

        return jsonify({
            "status": "success",
            "message": "Recording stopped",
            "filename": filename,
            "point_count": point_count,
            "start_time": start_time,
            "end_time": datetime.now().isoformat()
        })


@app.route('/recording/status', methods=['GET'])
def get_recording_status():
    """Get current recording status"""
    with recording_lock:
        return jsonify({
            "is_recording": recording_state["is_recording"],
            "filename": recording_state["filename"],
            "start_time": recording_state["start_time"],
            "point_count": recording_state["point_count"]
        })


@app.route('/recordings/list', methods=['GET'])
def list_recordings():
    """List all saved recordings"""
    recordings = []
    try:
        for f in os.listdir(RECORDINGS_DIR):
            if f.endswith('.csv'):
                filepath = os.path.join(RECORDINGS_DIR, f)
                stat = os.stat(filepath)

                # Count rows (subtract 1 for header)
                with open(filepath, 'r', encoding='utf-8') as file:
                    row_count = sum(1 for _ in file) - 1

                recordings.append({
                    'filename': f,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'point_count': max(0, row_count)
                })
    except Exception as e:
        print(f"[WARN] Error listing recordings: {e}")

    # Sort by modified date (newest first)
    recordings.sort(key=lambda x: x['modified'], reverse=True)
    return jsonify(recordings)


@app.route('/recordings/download/<filename>')
def download_recording(filename):
    """Download a specific recording file"""
    # Security: sanitize filename
    safe_filename = os.path.basename(filename)
    if not safe_filename.endswith('.csv'):
        return "Invalid file type", 400

    filepath = os.path.join(RECORDINGS_DIR, safe_filename)
    if os.path.isfile(filepath):
        return send_from_directory(RECORDINGS_DIR, safe_filename, as_attachment=True)
    return "Recording not found", 404


@app.route('/recordings/delete/<filename>', methods=['DELETE'])
def delete_recording(filename):
    """Delete a specific recording file"""
    # Security: sanitize filename
    safe_filename = os.path.basename(filename)
    if not safe_filename.endswith('.csv'):
        return jsonify({"status": "error", "message": "Invalid file type"}), 400

    filepath = os.path.join(RECORDINGS_DIR, safe_filename)
    if not os.path.isfile(filepath):
        return jsonify({"status": "error", "message": "Recording not found"}), 404

    try:
        os.remove(filepath)
        return jsonify({"status": "success", "message": f"Deleted {safe_filename}"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/missions/list', methods=['GET'])
def list_missions():
    """List all available mission CSV files from logs/ and logs/previous_data/"""
    missions = []

    # Get files from main logs directory
    try:
        for f in os.listdir(LOG_DIR):
            if f.endswith('.csv'):
                filepath = os.path.join(LOG_DIR, f)
                stat = os.stat(filepath)
                missions.append({
                    'filename': f,
                    'path': 'logs',
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
    except Exception as e:
        print(f"[WARN] Error reading logs directory: {e}")

    # Get files from archive directory
    try:
        for f in os.listdir(ARCHIVE_DIR):
            if f.endswith('.csv'):
                filepath = os.path.join(ARCHIVE_DIR, f)
                stat = os.stat(filepath)
                missions.append({
                    'filename': f,
                    'path': 'logs/previous_data',
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
    except Exception as e:
        print(f"[WARN] Error reading archive directory: {e}")

    # Get files from recordings directory
    try:
        for f in os.listdir(RECORDINGS_DIR):
            if f.endswith('.csv'):
                filepath = os.path.join(RECORDINGS_DIR, f)
                stat = os.stat(filepath)
                missions.append({
                    'filename': f,
                    'path': 'logs/recordings',
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'is_recording': True
                })
    except Exception as e:
        print(f"[WARN] Error reading recordings directory: {e}")

    # Sort by modified date (newest first)
    missions.sort(key=lambda x: x['modified'], reverse=True)

    return jsonify(missions)


@app.route('/api/missions/load/<path:filepath>', methods=['GET'])
def load_mission(filepath):
    """Load and parse a mission CSV file, returning data as JSON"""
    import csv
    import re

    # Security: only allow files from logs/, logs/previous_data/, or logs/recordings/
    if filepath.startswith('logs/recordings/'):
        full_path = os.path.join(RECORDINGS_DIR, os.path.basename(filepath))
    elif filepath.startswith('logs/previous_data/'):
        full_path = os.path.join(ARCHIVE_DIR, os.path.basename(filepath))
    elif filepath.startswith('logs/'):
        full_path = os.path.join(LOG_DIR, os.path.basename(filepath))
    else:
        return jsonify({'error': 'Invalid path'}), 400

    if not os.path.isfile(full_path):
        return jsonify({'error': 'File not found'}), 404

    try:
        data = {}
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        file_content = None

        for encoding in encodings:
            try:
                with open(full_path, 'r', encoding=encoding) as f:
                    file_content = f.read()
                break
            except UnicodeDecodeError:
                continue

        if file_content is None:
            return jsonify({'error': 'Could not decode file with any supported encoding'}), 500

        lines = file_content.strip().split('\n')
        if not lines:
            return jsonify({'error': 'Empty file'}), 400

        first_line = lines[0]

        # Detect format: comma-delimited, semicolon-delimited, or fixed-width
        is_semicolon_csv = ';' in first_line and first_line.count(';') >= 5
        is_comma_csv = ',' in first_line and first_line.count(',') >= 5

        if is_semicolon_csv:
            # Semicolon-delimited CSV (common in European Excel)
            import io
            reader = csv.DictReader(io.StringIO(file_content), delimiter=';')
            headers = reader.fieldnames

            for header in headers:
                data[header] = []

            for row in reader:
                for header in headers:
                    try:
                        value = float(row[header])
                    except (ValueError, TypeError):
                        value = row[header]
                    data[header].append(value)
        elif is_comma_csv:
            # Standard comma-delimited CSV
            import io
            reader = csv.DictReader(io.StringIO(file_content), delimiter=',')
            headers = reader.fieldnames

            for header in headers:
                data[header] = []

            for row in reader:
                for header in headers:
                    try:
                        value = float(row[header])
                    except (ValueError, TypeError):
                        value = row[header]
                    data[header].append(value)
        else:
            # Fixed-width format from legacy lorareceivertest.py
            # The header has known column names - use them to define expected structure
            expected_headers = [
                "Timestamp (MET)", "Elapsed [s]", "Temp_BME280 [°C]", "Hum [%]",
                "Press [hPa]", "Alt [m]", "Acc x [m/s²]", "Acc y [m/s²]",
                "Acc z [m/s²]", "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]",
                "Temp_MPU [°C]"
            ]

            # Check if this looks like our legacy format
            if "Timestamp (MET)" in first_line or "Elapsed [s]" in first_line:
                headers = expected_headers
            else:
                # Fallback to splitting by multiple spaces
                header_parts = re.split(r'\s{2,}', first_line.strip())
                headers = [h.strip() for h in header_parts if h.strip()]

            for header in headers:
                data[header] = []

            # Parse data rows - split by multiple whitespace
            for line in lines[1:]:
                if not line.strip():
                    continue
                # Skip MIN/MAX summary lines at the end
                if line.strip().startswith('MIN,') or line.strip().startswith('MAX,'):
                    continue

                # Split by multiple spaces (2+ spaces)
                values = re.split(r'\s{2,}', line.strip())
                values = [v.strip() for v in values if v.strip()]

                # Match values to headers
                for i, header in enumerate(headers):
                    if i < len(values):
                        try:
                            value = float(values[i])
                        except (ValueError, TypeError):
                            value = values[i]
                        data[header].append(value)

        # Get file info
        stat = os.stat(full_path)

        return jsonify({
            'filename': os.path.basename(full_path),
            'headers': headers,
            'data': data,
            'rowCount': len(data.get(headers[0], [])) if headers else 0,
            'fileSize': stat.st_size,
            'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
        })

    except Exception as e:
        return jsonify({'error': f'Failed to parse CSV: {str(e)}'}), 500


@app.route('/open_logs_folder', methods=['POST'])
def open_logs_folder():
    try:
        # Get absolute path to logs directory for security
        logs_path = os.path.abspath(LOG_DIR)

        # Ensure logs directory exists
        if not os.path.exists(logs_path):
            return jsonify({"status": "error", "message": "Logs directory does not exist"}), 404

        # Open folder based on operating system
        system = platform.system()
        if system == "Windows":
            os.startfile(logs_path)
        elif system == "Darwin":  # macOS
            subprocess.run(["open", logs_path], check=True)
        elif system == "Linux":
            subprocess.run(["xdg-open", logs_path], check=True)
        else:
            return jsonify({"status": "error", "message": f"Unsupported operating system: {system}"}), 400

        return jsonify({"status": "success", "message": "Logs folder opened successfully"})

    except Exception as e:
        return jsonify({"status": "error", "message": f"Failed to open logs folder: {str(e)}"}), 500


# ==================== Configuration API Endpoints ====================

@app.route('/config', methods=['GET'])
def get_config():
    """Get current configuration"""
    config = load_config()
    return jsonify(config)


@app.route('/config', methods=['POST'])
def update_config():
    """Update configuration"""
    try:
        new_config = request.json
        if not new_config:
            return jsonify({"error": "No configuration data provided"}), 400

        # Load current config and merge with new values
        current_config = load_config()

        # Update top-level values
        if 'sea_level_pressure' in new_config:
            current_config['sea_level_pressure'] = float(new_config['sea_level_pressure'])
        if 'update_frequency' in new_config:
            current_config['update_frequency'] = float(new_config['update_frequency'])
        if 'history_length' in new_config:
            current_config['history_length'] = int(new_config['history_length'])

        # Update transmission thresholds
        if 'transmission_thresholds' in new_config:
            for key, value in new_config['transmission_thresholds'].items():
                if key in current_config['transmission_thresholds']:
                    current_config['transmission_thresholds'][key] = float(value)

        save_config(current_config)
        return jsonify({"status": "success", "config": current_config})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/config/reset', methods=['POST'])
def reset_config():
    """Reset configuration to defaults"""
    try:
        save_config(DEFAULT_CONFIG.copy())
        return jsonify({"status": "success", "config": DEFAULT_CONFIG})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/config/profiles', methods=['GET'])
def get_profiles():
    """List all saved configuration profiles"""
    profiles = list_profiles()
    return jsonify({"profiles": profiles})


@app.route('/config/profiles/<name>', methods=['GET'])
def get_profile(name):
    """Load a specific configuration profile"""
    try:
        profile_path = get_profile_path(name)
        if not os.path.exists(profile_path):
            return jsonify({"error": f"Profile '{name}' not found"}), 404

        with open(profile_path, 'r', encoding='utf-8') as f:
            profile = json.load(f)
        return jsonify(profile)

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/config/profiles/<name>', methods=['POST'])
def save_profile(name):
    """Save current configuration as a named profile"""
    try:
        profile_path = get_profile_path(name)

        # Get current config or use provided config
        if request.json:
            config = request.json
        else:
            config = load_config()

        os.makedirs(CONFIG_DIR, exist_ok=True)
        with open(profile_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=4)

        return jsonify({"status": "success", "message": f"Profile '{name}' saved"})

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/config/profiles/<name>', methods=['DELETE'])
def delete_profile(name):
    """Delete a configuration profile"""
    try:
        profile_path = get_profile_path(name)
        if not os.path.exists(profile_path):
            return jsonify({"error": f"Profile '{name}' not found"}), 404

        os.remove(profile_path)
        return jsonify({"status": "success", "message": f"Profile '{name}' deleted"})

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/config/profiles/<name>/apply', methods=['POST'])
def apply_profile(name):
    """Apply a profile as the current configuration"""
    try:
        profile_path = get_profile_path(name)
        if not os.path.exists(profile_path):
            return jsonify({"error": f"Profile '{name}' not found"}), 404

        with open(profile_path, 'r', encoding='utf-8') as f:
            profile = json.load(f)

        save_config(profile)
        return jsonify({"status": "success", "config": profile})

    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ==================== Export API Endpoints ====================

def filter_history_by_time(hist, start=None, end=None):
    """Filter history data by time range (in seconds)"""
    if not hist.get("Elapsed [s]"):
        return hist

    elapsed = hist["Elapsed [s]"]
    if not elapsed:
        return hist

    # Determine indices within the time range
    indices = []
    for i, t in enumerate(elapsed):
        if t is None:
            continue
        if start is not None and t < start:
            continue
        if end is not None and t > end:
            continue
        indices.append(i)

    # Build filtered history
    filtered = {}
    for key, values in hist.items():
        filtered[key] = [values[i] for i in indices if i < len(values)]

    return filtered


@app.route('/export/json', methods=['GET'])
def export_json():
    """Export current session data as JSON with metadata"""
    global history

    start = request.args.get('start', type=float)
    end = request.args.get('end', type=float)

    # Filter data by time range if specified
    export_data = filter_history_by_time(history, start, end)

    # Build metadata
    elapsed = export_data.get("Elapsed [s]", [])
    metadata = {
        "export_timestamp": datetime.now().isoformat(),
        "project": "TRITON",
        "description": "Autonomous submarine sensor data",
        "point_count": len(elapsed),
        "time_range": {
            "start": min(elapsed) if elapsed else None,
            "end": max(elapsed) if elapsed else None,
            "filter_applied": {
                "start": start,
                "end": end
            } if start is not None or end is not None else None
        },
        "sensors": {
            "BME280": ["Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]"],
            "MPU6050": ["Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
                       "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]", "Temp_MPU [°C]"]
        }
    }

    # Compute basic statistics for each sensor
    statistics = {}
    for key, values in export_data.items():
        if key == "Elapsed [s]":
            continue
        numeric_values = [v for v in values if v is not None and isinstance(v, (int, float))]
        if numeric_values:
            statistics[key] = {
                "min": min(numeric_values),
                "max": max(numeric_values),
                "avg": sum(numeric_values) / len(numeric_values),
                "count": len(numeric_values)
            }

    output = {
        "metadata": metadata,
        "statistics": statistics,
        "data": export_data
    }

    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"triton_export_{timestamp}.json"

    response = Response(
        json.dumps(output, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )
    return response


@app.route('/export/excel', methods=['GET'])
def export_excel():
    """Export current session data as Excel (.xlsx) with multiple sheets"""
    global history

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    start = request.args.get('start', type=float)
    end = request.args.get('end', type=float)

    # Filter data by time range if specified
    export_data = filter_history_by_time(history, start, end)

    # Create workbook
    wb = openpyxl.Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: All Data
    ws_data = wb.active
    ws_data.title = "Sensor Data"

    # Headers
    headers = ["Elapsed [s]", "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]",
               "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
               "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]", "Temp_MPU [°C]"]

    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    num_rows = len(export_data.get("Elapsed [s]", []))
    for row_idx in range(num_rows):
        for col_idx, header in enumerate(headers, 1):
            values = export_data.get(header, [])
            value = values[row_idx] if row_idx < len(values) else None
            cell = ws_data.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, float):
                cell.number_format = '0.000'

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)

    # Sheet 2: BME280 Data
    ws_bme = wb.create_sheet("BME280")
    bme_headers = ["Elapsed [s]", "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]"]
    for col, header in enumerate(bme_headers, 1):
        cell = ws_bme.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx in range(num_rows):
        for col_idx, header in enumerate(bme_headers, 1):
            values = export_data.get(header, [])
            value = values[row_idx] if row_idx < len(values) else None
            cell = ws_bme.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, float):
                cell.number_format = '0.000'

    for col_idx, header in enumerate(bme_headers, 1):
        ws_bme.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)

    # Sheet 3: MPU6050 Data
    ws_mpu = wb.create_sheet("MPU6050")
    mpu_headers = ["Elapsed [s]", "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
                   "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]", "Temp_MPU [°C]"]
    for col, header in enumerate(mpu_headers, 1):
        cell = ws_mpu.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx in range(num_rows):
        for col_idx, header in enumerate(mpu_headers, 1):
            values = export_data.get(header, [])
            value = values[row_idx] if row_idx < len(values) else None
            cell = ws_mpu.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, float):
                cell.number_format = '0.000'

    for col_idx, header in enumerate(mpu_headers, 1):
        ws_mpu.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 12)

    # Sheet 4: Statistics
    ws_stats = wb.create_sheet("Statistics")
    stats_headers = ["Metric", "Min", "Max", "Average", "Count"]
    for col, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for key in headers[1:]:  # Skip Elapsed
        values = export_data.get(key, [])
        numeric_values = [v for v in values if v is not None and isinstance(v, (int, float))]
        if numeric_values:
            ws_stats.cell(row=row, column=1, value=key).border = thin_border
            ws_stats.cell(row=row, column=2, value=min(numeric_values)).border = thin_border
            ws_stats.cell(row=row, column=3, value=max(numeric_values)).border = thin_border
            ws_stats.cell(row=row, column=4, value=sum(numeric_values) / len(numeric_values)).border = thin_border
            ws_stats.cell(row=row, column=5, value=len(numeric_values)).border = thin_border
            for c in range(2, 5):
                ws_stats.cell(row=row, column=c).number_format = '0.000'
            row += 1

    for col_idx, header in enumerate(stats_headers, 1):
        ws_stats.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 2, 18)

    # Sheet 5: Metadata
    ws_meta = wb.create_sheet("Metadata")
    meta_data = [
        ("Project", "TRITON"),
        ("Description", "Autonomous submarine sensor data"),
        ("Export Time", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Data Points", num_rows),
        ("Time Range Start", min(export_data.get("Elapsed [s]", [0])) if export_data.get("Elapsed [s]") else "N/A"),
        ("Time Range End", max(export_data.get("Elapsed [s]", [0])) if export_data.get("Elapsed [s]") else "N/A"),
        ("Filter Start", start if start is not None else "None"),
        ("Filter End", end if end is not None else "None"),
    ]

    for row_idx, (label, value) in enumerate(meta_data, 1):
        cell_label = ws_meta.cell(row=row_idx, column=1, value=label)
        cell_label.font = Font(bold=True)
        cell_label.border = thin_border
        cell_value = ws_meta.cell(row=row_idx, column=2, value=value)
        cell_value.border = thin_border

    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 30

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"triton_export_{timestamp}.xlsx"

    return Response(
        buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@app.route('/export/pdf', methods=['GET'])
def export_pdf():
    """Export current session data as a formatted PDF report"""
    global history

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: pip install reportlab"}), 500

    start = request.args.get('start', type=float)
    end = request.args.get('end', type=float)

    # Filter data by time range if specified
    export_data = filter_history_by_time(history, start, end)

    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                           leftMargin=0.5*inch, rightMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=24,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        'Heading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
        textColor=colors.HexColor('#0066CC')
    )
    normal_style = styles['Normal']

    elements = []

    # Title
    elements.append(Paragraph("TRITON Sensor Data Report", title_style))
    elements.append(Spacer(1, 0.2*inch))

    # Metadata section
    elements.append(Paragraph("Report Information", heading_style))
    elapsed = export_data.get("Elapsed [s]", [])
    meta_data = [
        ["Project:", "TRITON - Autonomous Submarine Navigation System"],
        ["Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Data Points:", str(len(elapsed))],
        ["Time Range:", f"{min(elapsed):.2f}s - {max(elapsed):.2f}s" if elapsed else "No data"],
    ]
    if start is not None or end is not None:
        meta_data.append(["Filter Applied:", f"Start: {start if start else 'N/A'}, End: {end if end else 'N/A'}"])

    meta_table = Table(meta_data, colWidths=[1.5*inch, 5*inch])
    meta_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 0.3*inch))

    # Statistics section
    elements.append(Paragraph("Sensor Statistics", heading_style))

    headers = ["Elapsed [s]", "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]",
               "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
               "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]", "Temp_MPU [°C]"]

    stats_data = [["Metric", "Min", "Max", "Average", "Std Dev"]]
    for key in headers[1:]:  # Skip Elapsed
        values = export_data.get(key, [])
        numeric_values = [v for v in values if v is not None and isinstance(v, (int, float))]
        if numeric_values:
            avg = sum(numeric_values) / len(numeric_values)
            variance = sum((x - avg) ** 2 for x in numeric_values) / len(numeric_values)
            std_dev = variance ** 0.5
            stats_data.append([
                key,
                f"{min(numeric_values):.3f}",
                f"{max(numeric_values):.3f}",
                f"{avg:.3f}",
                f"{std_dev:.3f}"
            ])

    stats_table = Table(stats_data, colWidths=[2.5*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))

    # BME280 Data Table (first 50 rows as sample)
    elements.append(PageBreak())
    elements.append(Paragraph("BME280 Environmental Data (Sample)", heading_style))

    bme_headers = ["Elapsed [s]", "Temp [C]", "Humidity [%]", "Pressure [hPa]", "Altitude [m]"]
    bme_data = [bme_headers]
    num_rows = min(50, len(export_data.get("Elapsed [s]", [])))
    for i in range(num_rows):
        row = [
            f"{export_data.get('Elapsed [s]', [])[i]:.2f}" if i < len(export_data.get('Elapsed [s]', [])) else "",
            f"{export_data.get('Temp_BME280 [°C]', [])[i]:.2f}" if i < len(export_data.get('Temp_BME280 [°C]', [])) and export_data.get('Temp_BME280 [°C]', [])[i] is not None else "",
            f"{export_data.get('Hum [%]', [])[i]:.1f}" if i < len(export_data.get('Hum [%]', [])) and export_data.get('Hum [%]', [])[i] is not None else "",
            f"{export_data.get('Press [hPa]', [])[i]:.1f}" if i < len(export_data.get('Press [hPa]', [])) and export_data.get('Press [hPa]', [])[i] is not None else "",
            f"{export_data.get('Alt [m]', [])[i]:.1f}" if i < len(export_data.get('Alt [m]', [])) and export_data.get('Alt [m]', [])[i] is not None else "",
        ]
        bme_data.append(row)

    bme_table = Table(bme_data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1.4*inch, 1.2*inch])
    bme_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    elements.append(bme_table)

    if len(export_data.get("Elapsed [s]", [])) > 50:
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"<i>Showing first 50 of {len(export_data.get('Elapsed [s]', []))} data points. Use Excel or JSON export for complete data.</i>", normal_style))

    # MPU6050 Data Table (first 50 rows as sample)
    elements.append(PageBreak())
    elements.append(Paragraph("MPU6050 Inertial Data (Sample)", heading_style))

    mpu_headers = ["Time [s]", "Acc X", "Acc Y", "Acc Z", "Gyro X", "Gyro Y", "Gyro Z", "Temp [C]"]
    mpu_data = [mpu_headers]
    for i in range(num_rows):
        row = [
            f"{export_data.get('Elapsed [s]', [])[i]:.2f}" if i < len(export_data.get('Elapsed [s]', [])) else "",
            f"{export_data.get('Acc x [m/s²]', [])[i]:.3f}" if i < len(export_data.get('Acc x [m/s²]', [])) and export_data.get('Acc x [m/s²]', [])[i] is not None else "",
            f"{export_data.get('Acc y [m/s²]', [])[i]:.3f}" if i < len(export_data.get('Acc y [m/s²]', [])) and export_data.get('Acc y [m/s²]', [])[i] is not None else "",
            f"{export_data.get('Acc z [m/s²]', [])[i]:.3f}" if i < len(export_data.get('Acc z [m/s²]', [])) and export_data.get('Acc z [m/s²]', [])[i] is not None else "",
            f"{export_data.get('Gyro x [°/s]', [])[i]:.2f}" if i < len(export_data.get('Gyro x [°/s]', [])) and export_data.get('Gyro x [°/s]', [])[i] is not None else "",
            f"{export_data.get('Gyro y [°/s]', [])[i]:.2f}" if i < len(export_data.get('Gyro y [°/s]', [])) and export_data.get('Gyro y [°/s]', [])[i] is not None else "",
            f"{export_data.get('Gyro z [°/s]', [])[i]:.2f}" if i < len(export_data.get('Gyro z [°/s]', [])) and export_data.get('Gyro z [°/s]', [])[i] is not None else "",
            f"{export_data.get('Temp_MPU [°C]', [])[i]:.1f}" if i < len(export_data.get('Temp_MPU [°C]', [])) and export_data.get('Temp_MPU [°C]', [])[i] is not None else "",
        ]
        mpu_data.append(row)

    mpu_table = Table(mpu_data, colWidths=[0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch])
    mpu_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    elements.append(mpu_table)

    if len(export_data.get("Elapsed [s]", [])) > 50:
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"<i>Showing first 50 of {len(export_data.get('Elapsed [s]', []))} data points. Use Excel or JSON export for complete data.</i>", normal_style))

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"triton_report_{timestamp}.pdf"

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ==================== Motor Control API Endpoints ====================

import serial
import threading

# Motor control state
motor_state = {
    "throttle": 0,
    "target_throttle": 0,  # What we WANT the motor to be at
    "confirmed_throttle": 0,  # What the motor confirmed it's at
    "status": "disconnected",
    "last_command_time": None,
    "last_ack_time": None
}
motor_lock = threading.Lock()

# LoRa serial connection for commands (shared with receiver)
lora_serial = None
lora_lock = threading.Lock()

# Continuous transmission thread
motor_tx_thread = None
motor_tx_running = False

LORA_COM_PORT = 'COM8'
LORA_BAUD_RATE = 9600
MOTOR_TX_INTERVAL_FAST = 0.15  # When actively changing throttle
MOTOR_TX_INTERVAL_SLOW = 2.0   # When throttle is stable (allow sensor data through)
MOTOR_ACTIVE_DURATION = 3.0    # Stay in fast mode for 3 seconds after throttle change


def init_lora_serial():
    """Initialize LoRa serial connection for sending commands."""
    global lora_serial
    import sys
    try:
        if lora_serial is None or not lora_serial.is_open:
            lora_serial = serial.Serial(LORA_COM_PORT, LORA_BAUD_RATE, timeout=0.1)
            print(f"[INFO] LoRa serial initialized on {LORA_COM_PORT}", flush=True)
            return True
        else:
            return True  # Already open
    except Exception as e:
        print(f"[WARN] Could not initialize LoRa serial: {e}", flush=True)
        lora_serial = None  # Reset to ensure retry
        return False


def motor_transmit_loop():
    """Background thread that transmits throttle and receives sensor data."""
    global motor_state, motor_tx_running, lora_serial

    print("[MOTOR-TX] Transmission thread started", flush=True)

    last_throttle_change = 0
    last_target = 0

    while motor_tx_running:
        try:
            if init_lora_serial():
                with motor_lock:
                    target = motor_state["target_throttle"]
                    last_change_time = motor_state.get("last_command_time")

                # Detect throttle changes
                if target != last_target:
                    last_throttle_change = time.time()
                    last_target = target

                # Determine if we're in active mode (recent throttle change)
                time_since_change = time.time() - last_throttle_change
                is_active = time_since_change < MOTOR_ACTIVE_DURATION

                with lora_lock:
                    # Only send command if in active mode OR as periodic heartbeat
                    if is_active or time_since_change % MOTOR_TX_INTERVAL_SLOW < 0.2:
                        command = f"CMD:THROTTLE:{target}\n"
                        lora_serial.write(command.encode())
                        lora_serial.flush()

                    # Listen for incoming data (sensor data + ACKs)
                    # Longer listen window when not actively sending
                    listen_time = 0.1 if is_active else 0.5
                    listen_end = time.time() + listen_time

                    while time.time() < listen_end:
                        if lora_serial.in_waiting:
                            line = lora_serial.readline().decode(errors='ignore').strip()
                            if not line:
                                continue

                            print(f"[LORA-RAW] {line}", flush=True)

                            if line.startswith("ACK:"):
                                # Process motor acknowledgment
                                parts = line.split(":")
                                if len(parts) >= 4:
                                    try:
                                        actual = int(parts[2])
                                        with motor_lock:
                                            motor_state["confirmed_throttle"] = actual
                                            motor_state["throttle"] = actual
                                            motor_state["last_ack_time"] = datetime.now().isoformat()
                                            motor_state["status"] = "running" if actual > 0 else "stopped"
                                    except ValueError:
                                        pass
                            elif "," in line and not line.startswith("CMD:"):
                                # Process sensor data
                                print(f"[LORA-RX] Sensor data received", flush=True)
                                process_sensor_data(line)
                                with motor_lock:
                                    if motor_state["status"] == "disconnected":
                                        motor_state["status"] = "connected"
                        else:
                            time.sleep(0.02)  # Brief sleep if no data

                # Sleep interval depends on mode
                sleep_time = MOTOR_TX_INTERVAL_FAST if is_active else MOTOR_TX_INTERVAL_SLOW
                time.sleep(sleep_time)

        except Exception as e:
            print(f"[MOTOR-TX] Error: {e}")
            # Reset serial connection on error
            with lora_lock:
                try:
                    if lora_serial:
                        lora_serial.close()
                except:
                    pass
                lora_serial = None
            time.sleep(0.5)

    print("[MOTOR-TX] Continuous transmission thread stopped")


def start_motor_tx_thread():
    """Start the continuous motor command transmission thread."""
    global motor_tx_thread, motor_tx_running

    if motor_tx_thread and motor_tx_thread.is_alive():
        return

    motor_tx_running = True
    motor_tx_thread = threading.Thread(target=motor_transmit_loop, daemon=True)
    motor_tx_thread.start()


def stop_motor_tx_thread():
    """Stop the continuous motor command transmission thread."""
    global motor_tx_running
    motor_tx_running = False


def set_target_throttle(value):
    """Set the target throttle - will be continuously transmitted until confirmed."""
    global motor_state

    value = max(0, min(100, value))

    with motor_lock:
        motor_state["target_throttle"] = value
        motor_state["last_command_time"] = datetime.now().isoformat()

    print(f"[MOTOR] Target throttle set to {value}%")
    return True


# Legacy function for compatibility
def send_motor_command(command_type, value=0, wait_for_ack=True, max_retries=10, ack_timeout=2.0):
    """Send motor command - now uses continuous transmission approach."""
    if command_type == "THROTTLE":
        set_target_throttle(value)
        return True, "Target set"
    elif command_type in ["STOP", "ESTOP"]:
        set_target_throttle(0)
        return True, "Stop commanded"
    return False, "Unknown command"


# Start continuous transmission on module load
start_motor_tx_thread()


@app.route('/motor/status', methods=['GET'])
def get_motor_status():
    """Get current motor status."""
    with motor_lock:
        return jsonify(motor_state)


@app.route('/motor/throttle', methods=['POST'])
def set_motor_throttle():
    """Set motor throttle (0-100%)."""
    try:
        data = request.json
        throttle = int(data.get('throttle', 0))

        # Clamp throttle to valid range
        throttle = max(0, min(100, throttle))

        success, message = send_motor_command("THROTTLE", throttle)

        if success:
            return jsonify({"status": "success", "throttle": throttle, "message": message})
        else:
            return jsonify({"status": "error", "message": message}), 500

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/motor/stop', methods=['POST'])
def stop_motor():
    """Stop the motor (set throttle to 0)."""
    success, message = send_motor_command("STOP", 0)

    if success:
        return jsonify({"status": "success", "message": "Motor stopped"})
    else:
        return jsonify({"status": "error", "message": message}), 500


@app.route('/motor/emergency_stop', methods=['POST'])
def emergency_stop_motor():
    """Emergency stop - immediately cuts motor power."""
    success, message = send_motor_command("ESTOP", 0)

    if success:
        return jsonify({"status": "success", "message": "Emergency stop activated"})
    else:
        return jsonify({"status": "error", "message": message}), 500


@app.route('/motor/preset/<int:percent>', methods=['POST'])
def set_motor_preset(percent):
    """Set motor to preset throttle value."""
    # Validate preset
    if percent not in [0, 25, 50, 75, 100]:
        return jsonify({"status": "error", "message": "Invalid preset. Use 0, 25, 50, 75, or 100"}), 400

    success, message = send_motor_command("THROTTLE", percent)

    if success:
        return jsonify({"status": "success", "throttle": percent, "message": message})
    else:
        return jsonify({"status": "error", "message": message}), 500


# ==================== LoRa Receiver Thread ====================

LABELS = [
    "Elapsed [s]",
    "Temp_BME280 [°C]", "Hum [%]", "Press [hPa]", "Alt [m]",
    "Acc x [m/s²]", "Acc y [m/s²]", "Acc z [m/s²]",
    "Gyro x [°/s]", "Gyro y [°/s]", "Gyro z [°/s]",
    "Temp_MPU [°C]"
]
MIN_COLUMNS = len(LABELS) + 1

lora_receiver_running = False
lora_receiver_thread = None


def process_sensor_data(line):
    """Process incoming sensor data from Pi."""
    global latest_data, history

    fields = line.split(",")
    if len(fields) < MIN_COLUMNS:
        return False

    timestamp_str = fields[0]
    values = fields[1:]

    try:
        # Update latest_data
        data_dict = {}
        for i, label in enumerate(LABELS):
            if i < len(values):
                try:
                    data_dict[label] = float(values[i])
                except ValueError:
                    data_dict[label] = values[i]

        latest_data["timestamp"] = timestamp_str
        latest_data["data"] = data_dict

        # Update history for charts
        elapsed = float(values[0]) if values else 0
        history["Elapsed [s]"].append(elapsed)

        for key in history:
            if key != "Elapsed [s]":
                try:
                    idx = LABELS.index(key)
                    if idx < len(values):
                        history[key].append(float(values[idx]))
                    else:
                        history[key].append(None)
                except (ValueError, IndexError):
                    history[key].append(None)

        # Write to recording if active
        append_recording_data(data_dict, timestamp_str)

        return True

    except Exception as e:
        print(f"[LORA-RX] Error processing sensor data: {e}")
        return False


def process_ack(line):
    """Process acknowledgment from Pi motor controller."""
    global motor_state

    # Format: ACK:<type>:<value>:<status>
    parts = line.split(":")
    if len(parts) >= 4:
        cmd_type = parts[1]
        value = parts[2]
        status = parts[3]

        with motor_lock:
            motor_state["last_ack_time"] = datetime.now().isoformat()

            if status == "OK":
                print(f"[LORA-RX] ACK received: {cmd_type}={value} OK")
            else:
                print(f"[LORA-RX] ACK received: {cmd_type}={value} FAILED")


def lora_receiver_loop():
    """Background thread that receives data from LoRa."""
    global lora_receiver_running, motor_state, lora_serial

    print("[LORA-RX] Receiver thread started")

    while lora_receiver_running:
        try:
            if not init_lora_serial():
                time.sleep(2)
                continue

            with lora_lock:
                if lora_serial and lora_serial.is_open and lora_serial.in_waiting:
                    line = lora_serial.readline().decode(errors='ignore').strip()
                else:
                    line = None

            if line:
                print(f"[LORA-RX] {line}")

                # Check if it's an acknowledgment
                if line.startswith("ACK:"):
                    process_ack(line)

                # Check if it's sensor data (starts with timestamp)
                elif "," in line and not line.startswith("CMD:"):
                    if process_sensor_data(line):
                        with motor_lock:
                            if motor_state["status"] == "disconnected":
                                motor_state["status"] = "connected"

            time.sleep(0.05)  # Small delay to prevent CPU hogging

        except Exception as e:
            print(f"[LORA-RX] Error: {e}")
            # Reset serial connection on error
            with lora_lock:
                try:
                    if lora_serial:
                        lora_serial.close()
                except:
                    pass
                lora_serial = None
            time.sleep(1)

    print("[LORA-RX] Receiver thread stopped")


def start_lora_receiver():
    """Start the LoRa receiver background thread."""
    global lora_receiver_running, lora_receiver_thread

    if lora_receiver_thread and lora_receiver_thread.is_alive():
        print("[LORA-RX] Receiver already running")
        return

    lora_receiver_running = True
    lora_receiver_thread = threading.Thread(target=lora_receiver_loop, daemon=True)
    lora_receiver_thread.start()
    print("[LORA-RX] Started LoRa receiver thread")


def stop_lora_receiver():
    """Stop the LoRa receiver background thread."""
    global lora_receiver_running

    lora_receiver_running = False
    print("[LORA-RX] Stopping receiver thread...")


@app.route('/lora/status', methods=['GET'])
def get_lora_status():
    """Get LoRa connection status."""
    connected = lora_serial is not None and lora_serial.is_open
    return jsonify({
        "connected": connected,
        "port": LORA_COM_PORT,
        "baud": LORA_BAUD_RATE,
        "receiver_running": lora_receiver_running
    })


@app.route('/lora/start', methods=['POST'])
def start_lora():
    """Start LoRa receiver."""
    start_lora_receiver()
    return jsonify({"status": "success", "message": "LoRa receiver started"})


@app.route('/lora/stop', methods=['POST'])
def stop_lora():
    """Stop LoRa receiver."""
    stop_lora_receiver()
    return jsonify({"status": "success", "message": "LoRa receiver stopped"})


def cleanup():
    global lora_serial

    # Stop LoRa receiver thread
    stop_lora_receiver()

    # Close LoRa serial if open
    try:
        if lora_serial and lora_serial.is_open:
            lora_serial.close()
            print("[INFO] LoRa serial closed")
    except Exception as e:
        print(f"[WARN] Error closing LoRa serial: {e}")

    try:
        if platform.system() == "Windows":
            # Windows cleanup
            result = subprocess.check_output('netstat -ano | findstr :5000', shell=True).decode().strip()
            if result:
                lines = result.splitlines()
                for line in lines:
                    if 'LISTENING' in line:
                        pid = line.split()[-1]
                        try:
                            subprocess.run(f'taskkill /F /PID {pid}', shell=True, check=True)
                            print(f"[INFO] Killed leftover process on port 5000 (PID {pid})")
                        except subprocess.CalledProcessError:
                            pass  # Process may have already ended
        else:
            # Unix/Linux cleanup
            result = subprocess.check_output(["lsof", "-t", "-i:5000"]).decode().strip()
            if result:
                for pid in result.splitlines():
                    os.kill(int(pid), signal.SIGKILL)
                    print(f"[INFO] Killed leftover process on port 5000 (PID {pid})")
    except Exception as e:
        print(f"[WARN] Cleanup failed or port already free: {e}")

cleanup()
atexit.register(cleanup)


if __name__ == "__main__":
    import logging
    log = logging.getLogger('werkzeug')
    log.setLevel(logging.ERROR)  # Or use logging.CRITICAL to suppress all output

    # Start LoRa receiver thread for sensor data and motor acknowledgments
    print("[INFO] Starting LoRa receiver for motor control and sensor data...")
    start_lora_receiver()

    try:
        app.run(debug=False, host="0.0.0.0", use_reloader=False)
    except KeyboardInterrupt:
        print("\nCaught Ctrl+C, shutting down...")
        cleanup()
